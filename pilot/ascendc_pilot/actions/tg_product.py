# -*- coding: utf-8 -*-
"""Deterministic TG engines for init.yaml / plan.md / worklog.md / cases."""

from __future__ import annotations

import csv
import hashlib
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml

from testcase_agent import isolation, plan_packet, products, test_repo
from testcase_agent.init_status import InitGateError


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _tg(project_root: Path, ctx: dict[str, Any] | None = None) -> Path:
    from ascendc_pilot.paths import tg_root

    arch = str((ctx or {}).get("architecture") or "").strip() or None
    return tg_root(project_root, arch=arch)


def _run_id(ctx: dict[str, Any]) -> str:
    return str(ctx.get("run_id") or "").strip()


def _action_dir(project_root: Path, ctx: dict[str, Any], action_id: str) -> Path:
    from ascendc_pilot.paths import agent_root

    arch = str(ctx.get("architecture") or "").strip() or None
    rid = _run_id(ctx)
    return agent_root(project_root, arch) / "runs" / rid / "actions" / action_id


def _dump_yaml(path: Path, doc: Any) -> Path:
    isolation.assert_tg_write_path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    from testcase_agent.io import YAML_WIDTH

    path.write_text(
        yaml.safe_dump(doc, allow_unicode=True, sort_keys=False, width=YAML_WIDTH),
        encoding="utf-8",
    )
    return path


def _load_yaml(path: Path) -> Any:
    if not path.is_file():
        return {}
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return data


def _receipt(project_root: Path, ctx: dict[str, Any], name: str, payload: dict[str, Any]) -> Path:
    from ascendc_pilot.runs import receipts_dir

    body = dict(payload)
    body.setdefault("kind", "receipt")
    body.setdefault("written_at", _now())
    rid = _run_id(ctx)
    if rid:
        body.setdefault("run_id", rid)
    for key in ("workflow_id", "action_id", "architecture"):
        if ctx.get(key) and key not in body:
            body[key] = ctx[key]
    out = receipts_dir(project_root, rid or None) / name
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(yaml.safe_dump(body, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return out


_DEFAULT_INPUT_MARKERS = frozenset(
    {
        "no_repo_uo_query",
        "default_input",
        "none",
        "null",
        "-",
        "__default_input__",
    }
)


def _choice_from_raw(raw: str, project_root: Path | None = None) -> tuple[str, str]:
    from ascendc_pilot.human_interaction import (
        adopt_test_script_root,
        extract_existing_directory,
        extract_harness_git_url,
    )

    text = str(raw or "").strip()
    if text.lower() in _DEFAULT_INPUT_MARKERS:
        return "default_input", ""
    if text.lower() in {"have_repo", "custom", "stop"}:
        return "unset", ""
    extracted = extract_existing_directory(text) if text else ""
    if extracted:
        return "script_repo", extracted
    git_url = extract_harness_git_url(text) if text else ""
    if git_url and project_root is not None:
        adopted = adopt_test_script_root(project_root, git_url)
        stored = str(adopted.get("test_script_root") or "").strip()
        if adopted.get("ok") and stored:
            return "script_repo", stored
        err = str(adopted.get("message_zh") or adopted.get("error") or "HARNESS_CLONE_FAILED")
        return "clone_failed", err
    if text:
        try:
            path = Path(text).expanduser()
            if path.is_dir():
                return "script_repo", str(path.resolve())
        except OSError:
            pass
    return "unset", ""


def _test_script_root(project_root: Path, ctx: dict[str, Any]) -> str:
    from ascendc_pilot.actions.engines import _resolve_tg_ctx
    from ascendc_pilot.human_interaction import resolved_test_script_root

    tg_ctx = _resolve_tg_ctx(project_root, ctx)
    raw = str(tg_ctx.get("test_script_root") or ctx.get("test_script_root") or "").strip()
    raw = resolved_test_script_root(project_root, raw)
    if raw.lower() in _DEFAULT_INPUT_MARKERS or raw.lower() in {"have_repo", "custom", "stop"}:
        return ""
    return raw


def _harness_choice(project_root: Path, ctx: dict[str, Any]) -> tuple[str, str]:
    from ascendc_pilot.actions.engines import _resolve_tg_ctx
    from ascendc_pilot.human_interaction import (
        adopt_test_script_root,
        extract_existing_directory,
        is_in_tree_operator_tests,
        resolved_test_script_root,
    )
    from ascendc_pilot.state import load_state

    tg_ctx = _resolve_tg_ctx(project_root, ctx)
    picked = str(tg_ctx.get("test_script_root") or ctx.get("test_script_root") or "").strip()
    st = load_state(project_root) or {}
    live = str(st.get("workflow_id") or "") == "tg-init"
    if live:
        raw = resolved_test_script_root(project_root, picked)
        if st.get("test_script_confirmed"):
            return _choice_from_raw(raw, project_root)
        # Unconfirmed live run: operator-external directory or git URL is a user fact.
        # In-tree tests/ and default_input markers from pack overlay are guesses.
        if raw.lower() in _DEFAULT_INPUT_MARKERS or raw.lower() in {"have_repo", "custom", "stop"}:
            return "unset", ""
        extracted = extract_existing_directory(raw) if raw else ""
        if extracted and not is_in_tree_operator_tests(project_root, extracted):
            adopt_test_script_root(project_root, extracted)
            return "script_repo", extracted
        return _choice_from_raw(raw, project_root)
    return _choice_from_raw(picked, project_root)


def _discover_in_tree_tests(project_root: Path) -> str:
    for name in ("tests", "test"):
        candidate = Path(project_root) / name
        if candidate.is_dir():
            return str(candidate.resolve())
    return ""


def _uo_identity(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    from testcase_agent import product_uo

    from ascendc_pilot.actions.engines import _resolve_tg_ctx

    tg_ctx = _resolve_tg_ctx(project_root, ctx)
    return product_uo.identity(
        project_root,
        op_name=str(tg_ctx.get("op_name") or ""),
        architecture=str(tg_ctx.get("architecture") or ""),
    )


def _legal_key_count(project_root: Path, ctx: dict[str, Any]) -> int:
    from testcase_agent import product_uo

    from ascendc_pilot.actions.engines import _resolve_tg_ctx

    tg_ctx = _resolve_tg_ctx(project_root, ctx)
    try:
        rows = product_uo.legal_key_rows(
            project_root,
            op_name=str(tg_ctx.get("op_name") or ""),
            architecture=str(tg_ctx.get("architecture") or ""),
        )
    except Exception:
        return 0
    return len(rows)


def _collect_staging_mapping(root: Path) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    staging = _load_yaml(root / "staging.yaml")
    if isinstance(staging, dict):
        merged.update(staging)
    for part in sorted((root / "parts").glob("*.yaml")) if (root / "parts").is_dir() else []:
        doc = _load_yaml(part)
        if isinstance(doc, dict):
            for key, value in doc.items():
                if key not in merged or merged[key] in (None, "", {}, []):
                    merged[key] = value
                elif isinstance(merged.get(key), dict) and isinstance(value, dict):
                    merged[key] = {**merged[key], **value}
    return merged


def _collect_staging_text(root: Path, *, names: tuple[str, ...] = ("plan.md", "worklog.md", "staging.md")) -> str:
    """Prefer nonempty parts/<name>, then nonempty action-root <name>, then parts/*.md."""
    parts = root / "parts"
    for name in names:
        path = parts / name
        if path.is_file() and path.stat().st_size > 0:
            return path.read_text(encoding="utf-8")
    for name in names:
        path = root / name
        if path.is_file() and path.stat().st_size > 0:
            return path.read_text(encoding="utf-8")
    staging = _load_yaml(root / "staging.yaml")
    if isinstance(staging, dict):
        for key in ("text", "markdown", "plan_md", "worklog_md"):
            if staging.get(key):
                return str(staging[key])
    if parts.is_dir():
        chunks: list[str] = []
        for part in sorted(parts.glob("*.md")):
            if part.stat().st_size > 0:
                chunks.append(part.read_text(encoding="utf-8"))
        if chunks:
            return "\n\n".join(chunks)
    return ""


def _captured(project_root: Path, ctx: dict[str, Any], action_id: str) -> dict[str, Any]:
    from ascendc_pilot.actions.runtime import _load_tg_captured

    return _load_tg_captured(project_root, _run_id(ctx), action_id)


def _captured_text(project_root: Path, ctx: dict[str, Any], action_id: str) -> str:
    row = _captured(project_root, ctx, action_id)
    text = str(row.get("text") or "")
    if text.strip():
        return text
    doc = row.get("doc")
    if isinstance(doc, dict) and doc:
        return yaml.safe_dump(doc, allow_unicode=True, sort_keys=False)
    return ""


def _parse_analyze_actions(text: str) -> dict[str, Any] | None:
    raw = str(text or "").strip()
    if not raw:
        return None
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:yaml)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
    try:
        data = yaml.safe_load(raw)
    except yaml.YAMLError:
        return None
    if not isinstance(data, dict):
        return None
    if any(key in data for key in ("actions", "proof_requests", "refinement")):
        return data
    return None


def _extract_proof_requests(parsed: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not isinstance(parsed, dict):
        return []
    raw = parsed.get("proof_requests")
    if raw is None and isinstance(parsed.get("actions"), dict):
        raw = parsed["actions"].get("proof_requests")
    if not isinstance(raw, list):
        return []
    return [row for row in raw if isinstance(row, dict)]


def _mark_not_applicable(
    project_root: Path,
    ctx: dict[str, Any],
    action_ids: list[str],
    *,
    reason: str,
) -> None:
    rid = _run_id(ctx)
    for aid in action_ids:
        path = _action_dir(project_root, ctx, aid) / "not_applicable.yaml"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            yaml.safe_dump(
                {
                    "status": "not_applicable",
                    "action_id": aid,
                    "run_id": rid,
                    "reason": reason,
                },
                allow_unicode=True,
                sort_keys=False,
            ),
            encoding="utf-8",
        )


def _parse_captured_docs(text: str) -> list[Any]:
    raw = str(text or "").strip()
    if not raw:
        return []
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:yaml)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
    try:
        docs = list(yaml.safe_load_all(raw))
    except yaml.YAMLError:
        return []
    return [doc for doc in docs if doc is not None]


_PLAN_PROSE_HEADINGS = ("测什么", "覆盖什么", "怎么判定")


def _has_plan_prose(text: str) -> bool:
    blob = str(text or "")
    return all(re.search(rf"^##\s*{re.escape(h)}\s*$", blob, re.MULTILINE) for h in _PLAN_PROSE_HEADINGS)


def _assemble_plan_md(prose: str, mapping: dict[str, Any]) -> str:
    body = yaml.safe_dump(mapping, allow_unicode=True, sort_keys=False).rstrip() + "\n"
    return f"{prose.rstrip()}\n\n```yaml\n{body}```\n"


def _plan_fail(error: str, *, ask: str, **extra: Any) -> dict[str, Any]:
    out: dict[str, Any] = {
        "ok": False,
        "engine": "plan_promote",
        "error": error,
        "reason_code": error,
        "ask": ask,
        "retryable": True,
        "failure_class": "format_transport",
    }
    out.update(extra)
    return out


def _parse_captured_mapping(text: str, doc: dict[str, Any] | None = None) -> dict[str, Any]:
    if isinstance(doc, dict) and doc:
        return dict(doc)
    raw = str(text or "").strip()
    if not raw:
        return {}
    import re

    matches = list(re.finditer(r"```ya?ml\s*\n(.*?)```", raw, re.DOTALL | re.IGNORECASE)) if "```" in raw else []
    candidates = [m.group(1) for m in matches] + [raw]
    for body in candidates:
        try:
            parsed = yaml.safe_load(body)
        except Exception:  # noqa: BLE001
            continue
        if isinstance(parsed, dict):
            return parsed
    return {}


def _replay_dir(project_root: Path, ctx: dict[str, Any] | None = None) -> Path:
    return _tg(project_root, ctx) / "replay"


def _evidence_proofs(project_root: Path, mapping: dict[str, Any]) -> list[dict[str, Any]]:
    proofs: list[dict[str, Any]] = []
    try:
        from ascendc_pilot.evidence_window import disk_window_proof, first_evidence_locator
    except Exception:  # noqa: BLE001
        return proofs
    for col, row in (mapping or {}).items():
        if not isinstance(row, dict):
            continue
        ev = str(row.get("evidence") or "").strip()
        loc = first_evidence_locator(ev)
        if not loc:
            continue
        rel, spec = loc
        try:
            proof = disk_window_proof(project_root, path=rel, lines=spec)
        except Exception as exc:  # noqa: BLE001
            proof = {
                "ok": False,
                "error": "missing_file",
                "path": rel,
                "message_zh": str(exc)[:300],
            }
        proofs.append({"column": col, "evidence": ev, **proof})
    return proofs


def _goal_wants_test_generation(project_root: Path) -> bool:
    try:
        from ascendc_pilot.human_confirm import _auto_goal_wants_tests
        from ascendc_pilot.user_goal import load_user_goal
    except Exception:  # noqa: BLE001
        return False
    if _auto_goal_wants_tests(project_root):
        return True
    goal = load_user_goal(project_root) or {}
    for item in goal.get("public_plan") or []:
        if not isinstance(item, dict):
            continue
        if str(item.get("id") or "") in {"bind_harness", "generate_cases", "validate_cases"}:
            return True
    return False


def run_repo_scan(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    from ascendc_pilot.human_interaction import (
        adopt_test_script_root,
        extract_existing_directory,
        load_pending,
        pending_field,
    )

    pending = load_pending(project_root)
    if str(pending.get("status") or "") == "answered":
        field = pending_field(pending)
        answered = str(pending.get("answered_value") or "").strip()
        if field == "test_script_root" or extract_existing_directory(answered):
            adopt_test_script_root(project_root, answered)
    kind, root = _harness_choice(project_root, ctx)
    if kind == "clone_failed":
        return {
            "ok": False,
            "engine": "repo_scan",
            "error": "HARNESS_CLONE_FAILED",
            "message_zh": root or "测试仓 git URL 克隆失败。",
            "needs_human_decision": False,
            "test_script_root": "",
        }
    if kind == "unset":
        discovered = _discover_in_tree_tests(project_root)
        options: list[dict[str, str]] = [
            {
                "label": "没有测试仓，由 Agent 按算子约束生成",
                "value": "no_repo_uo_query",
                "description": "不扫描测试仓，按算子 UO 约束手写 init（default_input）",
            },
        ]
        if discovered:
            options.append(
                {
                    "label": f"使用已发现的仓内目录 {discovered}",
                    "value": discovered,
                    "description": "只有点选此项才把算子仓 tests/ 当作 harness；禁止代答",
                },
            )
        options.append(
            {
                "label": "有外部测试仓，或其他想法：在下面输入",
                "value": "custom",
                "description": "输入本地绝对路径、git 仓 URL，或其它说明（会打断当前确认）",
            }
        )
        question = (
            "尚未确认 test_script_root。"
            "请选择：没有测试仓则由 Agent 生成；"
            "确认仓内 tests/（若已发现）；"
            "或在最后一项输入外部仓路径 / git URL / 其它想法。"
            "未点选仓内项不得把 tests/ 当作 harness；不要静默 default_input。"
        )
        ask = {
            "header": "选择测试仓",
            "question": question,
            "prompt": question,
            "options": options,
            "allow_free_text": True,
            "field": "test_script_root",
        }
        return {
            "ok": False,
            "engine": "repo_scan",
            "needs_human_decision": True,
            "ask_question": ask,
            "message_zh": ask["question"],
            "test_script_root": "",
            "discovered_in_tree_tests": discovered,
        }
    inventory = test_repo.scan(root or None)
    ident = {}
    try:
        ident = _uo_identity(project_root, ctx)
    except Exception as exc:  # noqa: BLE001
        ident = {"error": str(exc)[:200]}
    declared = _legal_key_count(project_root, ctx)
    contract = test_repo.contract_from_inventory(inventory)
    doc = {
        "schema": "tg-repo-scan/v1",
        "ok": not str(inventory.get("error") or ""),
        "kind": kind,
        "test_script_root": root,
        "inventory": inventory,
        "contract": contract,
        "uo_digest": str(ident.get("sha256") or ident.get("digest") or ""),
        "uo_product": str(ident.get("path") or ""),
        "declared_key_count": declared,
        "declared_source": "product_uo.legal_key_rows",
    }
    out = _receipt(project_root, ctx, "repo_scan.yaml", doc)
    try:
        from testcase_agent.bind_parts import emit_bind_parts

        parts = _action_dir(project_root, ctx, "bind_init") / "parts"
        session = {
            "run_id": _run_id(ctx),
            "workflow_id": str(ctx.get("workflow_id") or "tg-init"),
            "phase": str(ctx.get("phase") or "bind"),
            "action_id": "bind_init",
            "actor_id": str(ctx.get("actor_id") or ""),
            "role_id": str(ctx.get("role_id") or ""),
            "action_session_id": str(ctx.get("action_session_id") or ""),
            "lease_id": str(ctx.get("lease_id") or ""),
            "prepare_nonce": str(ctx.get("prepare_nonce") or ""),
            "execution_mode": str(ctx.get("execution_mode") or "deterministic"),
        }
        identity = {"run_id": session["run_id"], "workflow_id": session["workflow_id"], "action_id": "bind_init"}
        try:
            from ascendc_pilot.ownership import artifact_identity_from_session

            identity = artifact_identity_from_session(session)
        except Exception:  # noqa: BLE001
            identity.setdefault("produced_by", "pilot-finalizer")
        emitted = emit_bind_parts(parts, scan=doc, identity=identity)
        doc["bind_parts"] = emitted
        out = _receipt(project_root, ctx, "repo_scan.yaml", doc)
    except Exception as exc:  # noqa: BLE001
        doc["bind_parts_error"] = str(exc)[:300]
        out = _receipt(project_root, ctx, "repo_scan.yaml", doc)
    return {"ok": True, "engine": "repo_scan", "artifact": out.as_posix(), **doc}


def run_bind_promote(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    tg = _tg(project_root, ctx)
    action_root = _action_dir(project_root, ctx, "bind_init")
    review_root = _action_dir(project_root, ctx, "bind_review")
    verdict = _load_yaml(review_root / "verdict.yaml")
    if verdict.get("ok") is not True:
        receipt = _receipt(
            project_root,
            ctx,
            "bind_promote.yaml",
            {"ok": False, "error": "REFEREE_REJECTED", "verdict": verdict},
        )
        return {
            "ok": False,
            "engine": "bind_promote",
            "error": "REFEREE_REJECTED",
            "message_zh": "没有引擎 verdict ok: true，禁止写入 tg/init.yaml。",
            "receipt": receipt.as_posix(),
        }
    from ascendc_pilot.yaml_check import format_yaml_error_zh, parse_yaml_mapping

    harness_path = action_root / "parts" / "harness.yaml"
    bind_path = action_root / "parts" / "bind.yaml"
    harness, harness_err = parse_yaml_mapping(harness_path)
    bind, bind_err = parse_yaml_mapping(bind_path)
    part_err = harness_err or bind_err
    if part_err:
        receipt = _receipt(
            project_root,
            ctx,
            "bind_promote.yaml",
            {"ok": False, **part_err},
        )
        return {
            "ok": False,
            "engine": "bind_promote",
            "error": str(part_err.get("error") or "BIND_PART_YAML_INVALID"),
            "reason_code": "BIND_PART_YAML_INVALID",
            "message_zh": format_yaml_error_zh(part_err),
            "path": part_err.get("path"),
            "line": part_err.get("line"),
            "column": part_err.get("column"),
            "receipt": receipt.as_posix(),
        }
    if not harness or not bind:
        receipt = _receipt(
            project_root,
            ctx,
            "bind_promote.yaml",
            {"ok": False, "error": "BIND_PARTS_MISSING", "harness": bool(harness), "bind": bool(bind)},
        )
        return {
            "ok": False,
            "engine": "bind_promote",
            "error": "BIND_PARTS_MISSING",
            "message_zh": "缺少 parts/harness.yaml 或 parts/bind.yaml，禁止 promote。",
            "receipt": receipt.as_posix(),
        }
    owned_bind = _load_yaml(action_root / "parts" / ".engine" / "bind.owned.yaml")
    owned_harness = _load_yaml(action_root / "parts" / ".engine" / "harness.owned.yaml")
    restore_err: list[Any] = []
    if owned_bind or owned_harness:
        from testcase_agent.bind_parts import restore_bind, restore_harness

        if owned_bind:
            bind, bind_restore_err = restore_bind(bind, owned_bind)
        else:
            bind_restore_err = []
        if owned_harness:
            harness, harness_restore_err = restore_harness(harness, owned_harness)
        else:
            harness_restore_err = []
        restore_err = list(bind_restore_err) + list(harness_restore_err)
        _dump_yaml(bind_path, bind)
        _dump_yaml(harness_path, harness)
    staged = _collect_staging_mapping(action_root)
    from ascendc_pilot.runs import receipts_dir

    scan = _load_yaml(receipts_dir(project_root, _run_id(ctx) or None) / "repo_scan.yaml")
    inventory = scan.get("inventory") if isinstance(scan.get("inventory"), dict) else {}
    contract = scan.get("contract") if isinstance(scan.get("contract"), dict) else test_repo.contract_from_inventory(inventory)
    ident = {}
    try:
        ident = _uo_identity(project_root, ctx)
    except Exception as exc:  # noqa: BLE001
        ident = {"error": str(exc)[:200]}

    scan_kind = str(scan.get("kind") or contract.get("kind") or "")
    kind = str(
        scan_kind
        or staged.get("kind")
        or bind.get("kind")
        or harness.get("kind")
        or ("script_repo" if inventory.get("kind") == "script_repo" else "default_input")
    )
    if not str(scan.get("test_script_root") or "").strip():
        kind = "default_input"
    table_kind = str(bind.get("table_kind") or staged.get("table_kind") or "csv").strip().lower()
    if table_kind not in {"csv", "xls", "xlsx"}:
        tables = inventory.get("tables") or []
        kinds = [str(t.get("kind") or "") for t in tables if isinstance(t, dict)]
        if "xlsx" in kinds:
            table_kind = "xlsx"
        elif "xls" in kinds:
            table_kind = "xls"
        else:
            table_kind = "csv"

    columns = bind.get("columns") or staged.get("columns") or contract.get("columns") or []
    if columns and isinstance(columns[0], str):
        columns = [{"name": c} for c in columns]

    findings: list[Any] = []
    for src in (harness.get("findings"), bind.get("findings"), staged.get("findings"), contract.get("findings")):
        if isinstance(src, list):
            findings.extend(src)
    findings.extend({"code": "bind_restore", "detail": item} for item in restore_err)

    mapping = products.mapping_as_dict(bind.get("mapping"))
    if not mapping:
        mapping = products.mapping_as_dict(staged.get("mapping"))
    domains = products.domains_as_dict(bind.get("domains") if bind.get("domains") is not None else bind.get("value_domains"))
    if not domains:
        domains = products.domains_as_dict(
            staged.get("domains") if staged.get("domains") is not None else staged.get("value_domains")
        )

    bind_call = bind.get("call") if isinstance(bind.get("call"), dict) else {}
    harness_call = harness.get("call") if isinstance(harness.get("call"), dict) else {}
    staged_call = staged.get("call") if isinstance(staged.get("call"), dict) else {}

    modes = dict(
        harness.get("modes") or staged.get("modes") or contract.get("modes") or {"precision": [], "perf": []}
    )
    modes.pop("candidates", None)
    doc = {
        "schema": products.INIT_SCHEMA,
        "kind": kind,
        "table_kind": table_kind,
        "entry": bind.get("entry") or staged.get("entry") or contract.get("entry") or "",
        "case_arg": bind.get("case_arg") or staged.get("case_arg") or contract.get("case_arg") or "",
        "modes": modes,
        "columns": columns,
        "defaults": bind.get("defaults") or staged.get("defaults") or contract.get("defaults") or {},
        "mapping": mapping,
        "domains": domains,
        "golden": harness.get("golden") or staged.get("golden") or {},
        "compare": harness.get("compare") or harness.get("script_compare") or staged.get("compare") or staged.get("script_compare") or {},
        "generate_inputs": harness.get("generate_inputs") or staged.get("generate_inputs") or {},
        "call": bind_call or harness_call or staged_call or {},
        "findings": findings,
        "test_script_root": str(scan.get("test_script_root") or _test_script_root(project_root, ctx)),
        "uo_digest": str(ident.get("sha256") or ident.get("digest") or scan.get("uo_digest") or ""),
        "uo_product": str(ident.get("path") or scan.get("uo_product") or ""),
        "declared_key_count": int(scan.get("declared_key_count") or 0),
        "declared_source": "product_uo.legal_key_rows",
        "confirmed": False,
        "project_root": Path(project_root).expanduser().resolve().as_posix(),
        "op_name": str(ctx.get("op_name") or ident.get("op_name") or Path(project_root).name),
        "architecture": str(ctx.get("architecture") or ident.get("architecture") or ""),
        "updated_at": _now(),
    }
    doc["findings"], doc["harness_capabilities"] = products.reconcile_findings(
        mapping,
        findings,
        generate_inputs=doc.get("generate_inputs"),
    )
    for key in ("precision_cmd", "perf_cmd", "corpus"):
        if harness.get(key) is not None:
            doc[key] = harness[key]
        elif staged.get(key) is not None:
            doc[key] = staged[key]
    # Merge dumps extracted structure. Validator owns bind legality
    # (including confirmed/uo.id). Primary judges whether a filled uo.id
    # names the right implementation symbol.
    path = products.dump_init(tg, doc)
    try:
        proofs = _evidence_proofs(project_root, mapping)
    except Exception as exc:  # noqa: BLE001 — proofs must not abort init.yaml
        proofs = [{"ok": False, "error": "evidence_proof_failed", "message_zh": str(exc)[:300]}]
    receipt = _receipt(
        project_root,
        ctx,
        "bind_promote.yaml",
        {"ok": True, "artifact": path.as_posix(), "evidence_proofs": proofs, "confirmed": False},
    )
    return {
        "ok": True,
        "engine": "bind_promote",
        "artifact": path.as_posix(),
        "receipt": receipt.as_posix(),
        "evidence_proofs": proofs,
        "confirmed": False,
    }


def run_validate_init(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    tg = _tg(project_root, ctx)
    try:
        doc = products.load_init(tg)
    except products.ProductError as exc:
        receipt = _receipt(project_root, ctx, "validate_init.yaml", {"ok": False, "error": str(exc)})
        return {
            "ok": False,
            "engine": "validate_init",
            "error": "INIT_INVALID",
            "reason_code": "INIT_INVALID",
            "ask": exc.ask,
            "message_zh": str(exc),
            "receipt": receipt.as_posix(),
        }
    errors = products.validate_init(doc)
    if errors:
        message_zh = "tg/init.yaml 校验失败：" + "；".join(str(item) for item in errors[:12])
        receipt = _receipt(
            project_root,
            ctx,
            "validate_init.yaml",
            {"ok": False, "error": "INIT_INVALID", "errors": errors},
        )
        return {
            "ok": False,
            "engine": "validate_init",
            "error": "INIT_INVALID",
            "reason_code": "INIT_INVALID",
            "errors": errors,
            "message_zh": message_zh,
            "receipt": receipt.as_posix(),
        }
    try:
        from testcase_agent.init_status import mark_init_confirmed

        mark_init_confirmed(tg, notes="Confirmed after Primary bind_review", project_root=project_root)
    except Exception as exc:  # noqa: BLE001
        receipt = _receipt(
            project_root,
            ctx,
            "validate_init.yaml",
            {"ok": False, "error": str(exc)[:300]},
        )
        return {
            "ok": False,
            "engine": "validate_init",
            "error": "INIT_INVALID",
            "reason_code": "INIT_INVALID",
            "message_zh": str(exc)[:400],
            "receipt": receipt.as_posix(),
        }
    receipt = _receipt(project_root, ctx, "validate_init.yaml", {"ok": True, "errors": []})
    return {"ok": True, "engine": "validate_init", "errors": [], "receipt": receipt.as_posix()}


def _compact_plan_scope_packet(
    project_root: Path,
    ctx: dict[str, Any],
    *,
    init_doc: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Prefetch from the Primary pin plus UO. Never git diff HEAD.

    The packet is what Plan Owner is allowed to consume, so the semantic half
    (observation vocabulary, changed observables, probeable locals, controls)
    is resolved here instead of being rediscovered from raw source.
    """
    arch = str(ctx.get("architecture") or "").strip()
    from ascendc_pilot.change_contract import (
        allow_legal_keys,
        changed_files_of,
        changed_hunks_of,
        compact_relevant_hunks,
        contract_public_meta,
        load_change_contract,
        scope_operator_hunks,
        verify_pinned_head,
    )

    contract = load_change_contract(project_root) or {}
    meta = contract_public_meta(contract)
    head_err = verify_pinned_head(project_root, contract)
    if head_err:
        packet = {
            "schema": plan_packet.PACKET_SCHEMA,
            "has_diff": False,
            "note": head_err.get("message_zh") or "HEAD mismatch",
            "changed_files": [],
            "change_contract": meta,
            "head_mismatch": head_err,
        }
        return packet
    files = changed_files_of(contract)
    scoped, extra = scope_operator_hunks(
        changed_hunks_of(contract),
        changed_files=files,
        operator_name=Path(project_root).name,
    )
    hunks = scoped
    relevant = compact_relevant_hunks(extra)
    has_diff = bool(files)
    note = ""
    if not has_diff:
        note = "无已 pin 的 change_contract.changed_files；不得把 git diff HEAD 当 PR 信号"
    intents = products.collect_intent_sources(project_root, architecture=arch)
    packet: dict[str, Any] = {
        "schema": plan_packet.PACKET_SCHEMA,
        "has_diff": has_diff,
        "note": note,
        "changed_files": files,
        "changed_hunks": hunks,
        "relevant_hunks": relevant,
        "change_contract": meta,
        "allow_legal_keys": allow_legal_keys(project_root),
        "identifiers": [],
        "ident_cards": [],
        "intent_sources": intents,
        "skip_around": True,
    }
    from ascendc_pilot.actions.engines import _resolve_tg_ctx

    tg_ctx = _resolve_tg_ctx(project_root, ctx)
    try:
        semantic = plan_packet.build_semantic_packet(
            project_root,
            op_name=str(tg_ctx.get("op_name") or Path(project_root).name),
            architecture=str(tg_ctx.get("architecture") or arch),
            changed_files=files,
            changed_hunks=hunks,
            relevant_hunks=extra,
            init_doc=init_doc,
            repo_root=_repo_root_for_contract(),
        )
    except Exception as exc:  # noqa: BLE001
        packet["semantic_packet_error"] = str(exc)[:300]
        packet["plan_route_card"] = plan_packet.build_plan_route_card(
            files, hunks, relevant_hunks=extra
        )
        return packet
    packet["identifiers"] = semantic.pop("identifiers", [])
    packet.update(semantic)
    return packet


def _repo_root_for_contract() -> Path:
    """Pilot checkout that owns the authoritative methodology files."""
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "skills").is_dir() and (parent / "prompts").is_dir():
            return parent
    return here.parents[3]


def run_plan_precheck(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    from testcase_agent.init_status import require_init_confirmed

    from ascendc_pilot.actions.engines import _resolve_tg_ctx

    tg_ctx = _resolve_tg_ctx(project_root, ctx)
    try:
        doc = require_init_confirmed(project_root, str(tg_ctx.get("op_name") or Path(project_root).name))
    except InitGateError as exc:
        return {"ok": False, "engine": "plan_precheck", "error": str(exc), "ask": exc.ask, "payload": exc.payload}
    from ascendc_pilot.change_contract import pr_change_gate

    gate = pr_change_gate(project_root)
    if gate:
        return gate
    from ascendc_pilot.change_contract import load_change_contract, verify_pinned_head

    head_err = verify_pinned_head(project_root, load_change_contract(project_root))
    if head_err:
        return head_err
    from ascendc_pilot.contract_sync import contract_drift_gate

    drift = contract_drift_gate()
    if drift:
        return drift
    declared = _legal_key_count(project_root, ctx)
    intents = products.collect_intent_sources(project_root, architecture=str(tg_ctx.get("architecture") or ""))
    packet = _compact_plan_scope_packet(project_root, ctx, init_doc=doc)
    packet_path = _receipt(project_root, ctx, "plan_scope_packet.yaml", packet)
    receipt = _receipt(
        project_root,
        ctx,
        "plan_precheck.yaml",
        {
            "ok": True,
            "declared_key_count": declared,
            "declared_source": "product_uo.legal_key_rows",
            "init_confirmed": True,
            "intent_sources": intents,
            "scope_packet": packet_path.as_posix(),
        },
    )
    return {
        "ok": True,
        "engine": "plan_precheck",
        "declared_key_count": declared,
        "receipt": receipt.as_posix(),
        "scope_packet": packet_path.as_posix(),
        "init": {"confirmed": True, "uo_digest": doc.get("uo_digest")},
    }


def run_plan_promote(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    ingest_cap = _captured(project_root, ctx, "plan_ingest")
    ingest_text = _captured_text(project_root, ctx, "plan_ingest")
    ingest_doc = ingest_cap.get("doc") if isinstance(ingest_cap.get("doc"), dict) else None
    mapping = _parse_captured_mapping(ingest_text, ingest_doc)
    if not mapping:
        return _plan_fail(
            "PLAN_INGEST_REQUIRED",
            ask="model",
            message_zh="缺少 Plan Owner YAML；下一发 `pilot_run(tg-plan)` intent 必须是 tg-plan/v3 全文，禁止空跑或让 promote 编一份。",
            rework_action_ids=["plan_ingest"],
        )
    schema = str(mapping.get("schema") or "").strip()
    if schema == "coverage-fragment/v1":
        return _plan_fail(
            "PLAN_INGEST_REQUIRED",
            ask="model",
            message_zh="plan_ingest 收到的是 fragment，不是 Owner 的 tg-plan/v3。先派 Plan Owner 汇总。",
            rework_action_ids=["plan_ingest"],
        )
    if schema and schema != products.PLAN_SCHEMA:
        return _plan_fail(
            "PLAN_INGEST_REQUIRED",
            ask="model",
            message_zh=f"plan_ingest schema {schema!r} 不是 {products.PLAN_SCHEMA}。",
            rework_action_ids=["plan_ingest"],
        )
    if not (mapping.get("targets") or []):
        return _plan_fail(
            "PLAN_INGEST_REQUIRED",
            ask="model",
            message_zh="Owner YAML 缺少 targets；禁止用 empty placeholder 继续 promote。",
            rework_action_ids=["plan_ingest"],
        )

    prose = products.render_plan_prose(mapping)
    text = _assemble_plan_md(prose, mapping)
    path = products.plan_path(_tg(project_root, ctx))
    isolation.assert_tg_write_path(path)
    path.write_text(text, encoding="utf-8")
    try:
        fence = products.parse_plan_fence(text)
    except products.ProductError as exc:
        return _plan_fail(
            "ENGINE_CONTRACT_VIOLATION",
            ask="promote",
            message_zh=str(exc),
            rework_action_ids=["plan_promote"],
        )
    return {
        "ok": True,
        "engine": "plan_promote",
        "artifact": path.as_posix(),
        "plan_hash": products.plan_hash(text),
        "target_count": len(fence.get("targets") or []),
    }


def _open_failure_reason(ledger: dict[str, Any]) -> str:
    from testcase_agent.coverage.contract import CASE_REFINABLE, CONTROL_GAP, OBSERVATION_GAP, PLAN_INVALID
    from testcase_agent.coverage.ledger import open_ids

    rows = ledger.get("obligations") if isinstance(ledger.get("obligations"), dict) else {}
    classes: list[str] = []
    for oid in open_ids(ledger):
        row = rows.get(oid) if isinstance(rows.get(oid), dict) else {}
        classes.append(str(row.get("failure_class") or CASE_REFINABLE).strip() or CASE_REFINABLE)
    for code in (PLAN_INVALID, CONTROL_GAP, OBSERVATION_GAP):
        if code in classes:
            return code
    if classes:
        return CASE_REFINABLE
    return ""


def _observe_fields(project_root: Path, ctx: dict[str, Any]) -> set[str] | None:
    try:
        from testcase_agent import product_uo
        from ascendc_pilot.actions.engines import _resolve_tg_ctx

        tg_ctx = _resolve_tg_ctx(project_root, ctx)
        return product_uo.replay_observe_fields(
            project_root,
            op_name=str(tg_ctx.get("op_name") or ""),
            architecture=str(tg_ctx.get("architecture") or ""),
        )
    except Exception:  # noqa: BLE001
        return None


def _init_mapping(init_doc: dict[str, Any]) -> dict[str, Any] | None:
    """None when init has no mapping — skip confirmed+active. Empty {} still checks."""
    if "mapping" not in init_doc or init_doc.get("mapping") is None:
        return None
    return products.mapping_as_dict(init_doc.get("mapping"))


def _primary_observations(project_root: Path, fence: dict[str, Any]) -> set[str] | None:
    obs: set[str] = set()
    req = fence.get("requirement") if isinstance(fence.get("requirement"), dict) else {}
    for raw in (
        req.get("observations")
        or req.get("primary_observations")
        or fence.get("primary_observations")
        or []
    ):
        name = str(raw or "").strip()
        if name:
            obs.add(name)
    try:
        from ascendc_pilot.change_contract import load_change_contract

        contract = load_change_contract(project_root) or {}
        for raw in contract.get("primary_observations") or contract.get("observations") or []:
            name = str(raw or "").strip()
            if name:
                obs.add(name)
    except Exception:  # noqa: BLE001
        pass
    return obs or None


def _load_scope_packet(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any] | None:
    """The packet plan_precheck wrote for this run, when it is still on disk."""
    try:
        from ascendc_pilot.runs import receipts_dir

        path = receipts_dir(project_root, str(ctx.get("run_id") or "")) / "plan_scope_packet.yaml"
        if not path.is_file():
            return None
        doc = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:  # noqa: BLE001
        return None
    return doc if isinstance(doc, dict) else None


def run_plan_validate(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    tg = _tg(project_root, ctx)
    try:
        init_doc = products.load_init(tg)
        text, fence = products.load_plan(tg)
    except products.ProductError as exc:
        return {"ok": False, "engine": "plan_validate", "error": str(exc), "ask": exc.ask}
    from ascendc_pilot.change_contract import allow_legal_keys

    semantic = products.semantic_plan_hash(fence)
    observations = _primary_observations(project_root, fence)
    errors = products.validate_plan_fence(
        fence,
        init_columns=products.column_names(init_doc),
        init_mapping=_init_mapping(init_doc),
        allow_legal_keys=allow_legal_keys(project_root),
        observe_fields=_observe_fields(project_root, ctx),
        primary_observations=observations,
    )
    errors.extend(products.validate_plan_prose(text, fence))
    from testcase_agent.coverage.contract import validate_against_packet

    errors.extend(validate_against_packet(fence, _load_scope_packet(project_root, ctx)))
    if str(fence.get("mode") or "").strip() in {"tilingkey_full_coverage", "T=D", "t_equals_d"}:
        errors.append("tilingkey_full_coverage / T=D is not a plan mode")
    ok = not errors
    receipt = _receipt(
        project_root,
        ctx,
        "plan_validate.yaml",
        {
            "ok": ok,
            "errors": errors,
            "plan_hash": products.plan_hash(text),
            "semantic_plan_hash": semantic,
            "test_harness_gap_pending": products.pending_test_harness_gap(text, fence),
        },
    )
    return {
        "ok": ok,
        "engine": "plan_validate",
        "errors": errors,
        "receipt": receipt.as_posix(),
        "reason_code": "" if ok else "PLAN_INVALID",
        "semantic_plan_hash": semantic,
    }


def run_solve_precheck(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    tg = _tg(project_root, ctx)
    try:
        text, fence = products.load_plan(tg)
        init_doc = products.load_init(tg)
    except products.ProductError as exc:
        return {"ok": False, "engine": "solve_precheck", "error": str(exc), "ask": exc.ask}
    if not products.is_plan_approved(fence):
        return {
            "ok": False,
            "engine": "solve_precheck",
            "error": "plan.md is not approved",
            "ask": "plan_required",
            "next": "/tg-plan",
            "reason_code": "PLAN_INVALID",
        }
    if products.pending_test_harness_gap(text, fence):
        return {
            "ok": False,
            "engine": "solve_precheck",
            "error": "test_harness_gap is pending; CE-apply the test-script repo then /tg-init",
            "ask": "test_harness_gap_pending",
            "reason_code": "HARNESS_CONTROL_GAP",
        }
    current = products.semantic_plan_hash(fence)
    stamped = str(fence.get("plan_hash") or "").strip()
    if stamped and stamped != current:
        return {
            "ok": False,
            "engine": "solve_precheck",
            "error": "plan.md changed after approve; re-run /tg-plan validate+approve",
            "ask": "plan_required",
            "next": "/tg-plan",
            "reason_code": "PLAN_INVALID",
        }
    from ascendc_pilot.change_contract import allow_legal_keys

    errors = products.validate_plan_fence(
        fence,
        init_columns=products.column_names(init_doc),
        init_mapping=_init_mapping(init_doc),
        allow_legal_keys=allow_legal_keys(project_root),
        observe_fields=_observe_fields(project_root, ctx),
    )
    if errors:
        return {
            "ok": False,
            "engine": "solve_precheck",
            "error": "approved plan failed executability contract",
            "errors": errors,
            "reason_code": "PLAN_INVALID",
        }
    receipt = _receipt(
        project_root,
        ctx,
        "solve_precheck.yaml",
        {"ok": True, "plan_hash": products.plan_hash(text), "semantic_plan_hash": current},
    )
    return {"ok": True, "engine": "solve_precheck", "receipt": receipt.as_posix()}


def run_compile_obligations(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    from testcase_agent.coverage.compile import PlanCompileError, compile_obligations
    from testcase_agent.coverage.contract import obligation_identity
    from testcase_agent.coverage.ledger import dump_worklog, parse_worklog_fence, seed_ledger

    tg = _tg(project_root, ctx)
    try:
        _text, fence = products.load_plan(tg)
        init_doc = products.load_init(tg)
    except products.ProductError as exc:
        return {"ok": False, "engine": "compile_obligations", "error": str(exc), "ask": exc.ask}
    from ascendc_pilot.change_contract import allow_legal_keys

    errors = products.validate_plan_fence(
        fence,
        init_columns=products.column_names(init_doc),
        init_mapping=_init_mapping(init_doc),
        allow_legal_keys=allow_legal_keys(project_root),
        observe_fields=_observe_fields(project_root, ctx),
    )
    if errors:
        return {
            "ok": False,
            "engine": "compile_obligations",
            "error": "PLAN_INVALID",
            "errors": errors,
            "reason_code": "PLAN_INVALID",
        }
    legal_keys = None
    cov = fence.get("coverage") if isinstance(fence.get("coverage"), dict) else {}
    if str(cov.get("enumerate") or "").strip() == "legal_keys":
        from testcase_agent import product_uo

        from ascendc_pilot.actions.engines import _resolve_tg_ctx

        tg_ctx = _resolve_tg_ctx(project_root, ctx)
        try:
            legal_keys = product_uo.legal_key_rows(
                project_root,
                op_name=str(tg_ctx.get("op_name") or ""),
                architecture=str(tg_ctx.get("architecture") or ""),
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "engine": "compile_obligations", "error": f"legal_keys unavailable: {exc}"}
    try:
        obligations = compile_obligations(fence, legal_keys=legal_keys)
    except PlanCompileError as exc:
        return {
            "ok": False,
            "engine": "compile_obligations",
            "error": "PLAN_INVALID",
            "errors": exc.errors,
            "reason_code": "PLAN_INVALID",
        }
    path = products.worklog_path(tg)
    existing = path.read_text(encoding="utf-8") if path.is_file() else ""
    old = parse_worklog_fence(existing)
    old_rows = old.get("obligations") if isinstance(old.get("obligations"), dict) else {}
    ledger = seed_ledger(obligations)
    for oid, row in (ledger.get("obligations") or {}).items():
        row["identity"] = obligation_identity(row)
        prev = old_rows.get(oid) if isinstance(old_rows.get(oid), dict) else {}
        if prev.get("status") and str(prev.get("identity") or "") == row["identity"]:
            row["status"] = prev.get("status")
            if prev.get("signature"):
                row["signature"] = prev.get("signature")
            if prev.get("witness"):
                row["witness"] = prev.get("witness")
    if isinstance(old.get("signatures"), list):
        ledger["signatures"] = list(old.get("signatures") or [])
    isolation.assert_tg_write_path(path)
    path.write_text(dump_worklog(ledger, prose=""), encoding="utf-8")
    try:
        from testcase_agent.plan_fill import ensure_v3
        from testcase_agent.solve_fill import index_plan

        indexed = index_plan(ensure_v3(fence, init_doc), init_doc)
        idx_path = tg / "solve_index.yaml"
        isolation.assert_tg_write_path(idx_path)
        _dump_yaml(
            idx_path,
            {
                "schema": "tg-solve-index/v1",
                "needs_hit": indexed.get("needs_hit") or [],
                "auto": indexed.get("auto") or [],
                "guards": indexed.get("guards") or [],
                "obligation_count": len(ledger.get("obligations") or {}),
            },
        )
    except Exception:  # noqa: BLE001
        idx_path = None
    return {
        "ok": True,
        "engine": "compile_obligations",
        "artifact": path.as_posix(),
        "count": len(ledger.get("obligations") or {}),
        "solve_index": None if idx_path is None else idx_path.as_posix(),
    }


def run_construct_promote(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    tg = _tg(project_root, ctx)
    init_doc = products.load_init(tg)
    captured = _captured(project_root, ctx, "construct_cases")
    staged = _parse_captured_mapping(str(captured.get("text") or ""), captured.get("doc") if isinstance(captured.get("doc"), dict) else None)
    if not staged:
        staged = _collect_staging_mapping(_action_dir(project_root, ctx, "construct_cases"))
    try:
        from testcase_agent.plan_fill import load_yaml
        from testcase_agent.solve_fill import assemble_solve, is_solve_fill

        if is_solve_fill(staged):
            _, plan_doc = products.load_plan(tg)
            staged = assemble_solve(staged, plan_doc, init_doc)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "engine": "construct_promote", "error": str(exc)}
    rows = staged.get("rows") or staged.get("cases") or []
    if rows and not isinstance(rows, list):
        return {"ok": False, "engine": "construct_promote", "error": "rows is not a list"}
    recipe = staged.get("recipe") if isinstance(staged.get("recipe"), dict) else {}
    columns = [c["name"] if isinstance(c, dict) else str(c) for c in (init_doc.get("columns") or [])]
    extra = staged.get("columns") or []
    for col in extra:
        name = col["name"] if isinstance(col, dict) else str(col)
        if name and name not in columns:
            columns.append(name)
    if not columns and rows and isinstance(rows[0], dict):
        columns = [str(k) for k in rows[0].keys()]
    replay = _replay_dir(project_root, ctx)
    replay.mkdir(parents=True, exist_ok=True)
    pending = {"columns": columns, "rows": [r for r in rows if isinstance(r, dict)], "recipe": recipe}
    if str(recipe.get("kind") or "") == "enumerate_legal_keys":
        pending["rows"] = _materialize_legal_key_batch(project_root, ctx, recipe, columns, init_doc)
    pending_path = replay / "pending.yaml"
    isolation.assert_tg_write_path(pending_path)
    _dump_yaml(pending_path, pending)
    receipt = _receipt(
        project_root,
        ctx,
        "construct_promote.yaml",
        {"ok": True, "rows": len(pending.get("rows") or []), "recipe": recipe.get("kind") or "", "pending": pending_path.as_posix()},
    )
    cases = products.cases_path(tg, str(init_doc.get("table_kind") or "csv"))
    return {
        "ok": True,
        "engine": "construct_promote",
        "receipt": receipt.as_posix(),
        "rows": len(pending.get("rows") or []),
        "wrote_cases": cases.is_file(),
        "pending": pending_path.as_posix(),
    }


def _materialize_legal_key_batch(
    project_root: Path,
    ctx: dict[str, Any],
    recipe: dict[str, Any],
    columns: list[str],
    init_doc: dict[str, Any],
) -> list[dict[str, str]]:
    from testcase_agent import product_uo
    from testcase_agent.coverage.ledger import parse_worklog_fence

    from ascendc_pilot.actions.engines import _resolve_tg_ctx

    tg_ctx = _resolve_tg_ctx(project_root, ctx)
    try:
        key_rows = product_uo.legal_key_rows(
            project_root,
            op_name=str(tg_ctx.get("op_name") or ""),
            architecture=str(tg_ctx.get("architecture") or ""),
        )
    except Exception:
        key_rows = []
    worklog = products.worklog_path(_tg(project_root, ctx))
    ledger = parse_worklog_fence(worklog.read_text(encoding="utf-8") if worklog.is_file() else "")
    open_keys: set[int] = set()
    for row in (ledger.get("obligations") or {}).values():
        if not isinstance(row, dict):
            continue
        if str(row.get("status") or "OPEN") != "OPEN":
            continue
        if row.get("tiling_key") is None:
            continue
        try:
            open_keys.add(int(row.get("tiling_key")))
        except (TypeError, ValueError):
            continue
    try:
        batch_size = int(recipe.get("batch_size") or 64)
    except (TypeError, ValueError):
        batch_size = 64
    batch_size = max(1, min(batch_size, 512))
    fillers = recipe.get("fillers") if isinstance(recipe.get("fillers"), dict) else {}
    defaults = {}
    for col in init_doc.get("columns") or []:
        if isinstance(col, dict) and col.get("name") is not None:
            defaults[str(col["name"])] = "" if col.get("default") is None else str(col.get("default"))
    out: list[dict[str, str]] = []
    for raw in key_rows:
        if not isinstance(raw, dict):
            continue
        try:
            key_i = int(raw.get("tiling_key") if raw.get("tiling_key") is not None else raw.get("key"))
        except (TypeError, ValueError):
            continue
        if open_keys and key_i not in open_keys:
            continue
        row = dict(defaults)
        for key, value in raw.items():
            row[str(key)] = "" if value is None else str(value)
        for key, value in fillers.items():
            row[str(key)] = "" if value is None else str(value)
        row.setdefault("tiling_key", str(key_i))
        out.append({c: row.get(c, "") for c in (columns or list(row.keys()))} if columns else row)
        if len(out) >= batch_size:
            break
    cache = _replay_dir(project_root, ctx)
    cache.mkdir(parents=True, exist_ok=True)
    products.write_cases_csv(cache / "batch.csv", columns or (list(out[0].keys()) if out else []), out)
    return out


def _read_cases(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.is_file():
        return [], []
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            cols = [str(c) for c in (reader.fieldnames or []) if c]
            rows = [{k: "" if v is None else str(v) for k, v in row.items()} for row in reader]
            return cols, rows
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            return [], []
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(c or "") for c in next(it, ())]
        rows = []
        for raw in it:
            rows.append({header[i]: "" if i >= len(raw) or raw[i] is None else str(raw[i]) for i in range(len(header))})
        return header, rows
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError:
            return [], []
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
        rows = []
        for r in range(1, sheet.nrows):
            rows.append({header[c]: str(sheet.cell_value(r, c)) for c in range(sheet.ncols)})
        return header, rows
    return [], []


def _live_replay(ctx: dict[str, Any]) -> bool:
    if "live_replay" in ctx:
        return bool(ctx.get("live_replay"))
    if str(os.environ.get("TG_CLOSURE_CI") or "").strip().lower() in {"1", "true", "yes"}:
        return False
    if str(os.environ.get("UO_OPERATOR") or "").startswith("_synthetic"):
        return False
    return True


_REPLAY_FAIL_ZH = {
    "WSL_UNAVAILABLE": "未检测到可用的 WSL 发行版，Host replay 无法启动，不能进入 analyze。",
    "WSL_DISTRO_AMBIGUOUS": "存在多个 WSL 发行版，请设置 UO_REPLAY_DISTRO 后再跑 /tg-solve。",
    "WSL_DISTRO_UNRESOLVED": "指定的 WSL 发行版不在 `wsl -l -q` 列表中。",
    "WSL_BUILD_DEPS_MISSING": "WSL 缺少 g++/cmake，非交互 apt-get 失败。",
    "WSL_PATH_FAILED": "无法把 Windows 路径映射进 WSL（wslpath 失败）。",
    "CANN_ENV_NOT_FOUND": "未找到 UO 解包的 CANN 树（UO_CANN_ROOT / _cann/pkg），Host replay 不能开始。",
    "CANN_ENV_NOT_READY": "CANN 环境未就绪，Host replay 不能开始。",
}


def _replay_bootstrap_failure(exc: BaseException) -> tuple[str, str]:
    text = str(exc or "")
    code = "REPLAY_BOOTSTRAP_FAILED"
    blob = text.replace("\\", "/")
    for token in (
        "WSL_UNAVAILABLE",
        "WSL_DISTRO_AMBIGUOUS",
        "WSL_DISTRO_UNRESOLVED",
        "WSL_BUILD_DEPS_MISSING",
        "WSL_PATH_FAILED",
        "CANN_ENV_NOT_FOUND",
        "CANN_ENV_NOT_READY",
    ):
        if token in blob:
            code = token
            break
    else:
        if blob.startswith("REPLAY_BOOTSTRAP_FAILED:"):
            inner = blob.split(":", 2)[1].strip()
            if inner:
                code = inner.split("/")[0].split(":")[0] or code
    message_zh = _REPLAY_FAIL_ZH.get(code, f"Host replay 环境未就绪（{code}），不能进入 analyze。")
    return code, message_zh


def _load_pending_rows(project_root: Path, ctx: dict[str, Any]) -> list[dict[str, str]]:
    pending_path = _replay_dir(project_root, ctx) / "pending.yaml"
    if pending_path.is_file():
        doc = _load_yaml(pending_path)
        if isinstance(doc, dict):
            rows = doc.get("rows") or []
            if isinstance(rows, list):
                return [{str(k): "" if v is None else str(v) for k, v in row.items()} for row in rows if isinstance(row, dict)]
    batch = _replay_dir(project_root, ctx) / "batch.csv"
    if batch.is_file():
        _cols, rows = _read_cases(batch)
        return rows
    return []


def _observe_from_verdict(row: dict[str, str], item: Any) -> dict[str, Any]:
    replay = {
        "tiling_key": getattr(item, "key", None) if item is not None else row.get("tiling_key"),
        "dims": dict(getattr(item, "dims", {}) or {}) if item is not None else {},
        "logged": dict(getattr(item, "logged", {}) or {}) if item is not None else {},
        "diag": dict(getattr(item, "diag", {}) or {}) if item is not None else {},
        "tiling_data": dict(getattr(item, "tiling_data", {}) or {}) if item is not None else {},
        "ok": bool(getattr(item, "ok", False)) if item is not None else False,
        "reject": str(getattr(item, "reject", "") or "") if item is not None else str(row.get("reject") or ""),
    }
    for blob in (replay["logged"], replay["diag"], replay["dims"], replay["tiling_data"]):
        if isinstance(blob, dict):
            replay.update({str(k): v for k, v in blob.items() if str(k) not in replay})
    probes = dict(getattr(item, "probes", {}) or {}) if item is not None else {}
    return {"case": dict(row), "replay": replay, "probe": probes}


def _judge_rows(rows: list[dict[str, str]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    from testcase_agent.closure.oracle import HostOracle

    class _Row:
        def __init__(self, tag: str, row: dict[str, str]) -> None:
            self.tag = tag
            self.row = row

    oracle = HostOracle()
    tagged = [_Row(str(row.get("Testcase_Name") or f"case_{i}"), row) for i, row in enumerate(rows)]
    judged = oracle.judge(tagged, tag="tg_solve")
    verdicts: list[dict[str, Any]] = []
    observes: list[dict[str, Any]] = []
    for i, item in enumerate(judged):
        row = rows[i] if i < len(rows) else {}
        observe = _observe_from_verdict(row, item)
        observes.append(observe)
        verdicts.append(
            {
                "case_id": item.case_id,
                "ok": item.ok,
                "tiling_key": item.key,
                "reject": item.reject,
                "judged": item.judged,
                "dims": item.dims,
                "logged": item.logged,
                "diag": item.diag,
                "probes": item.probes,
                "tiling_data": item.tiling_data,
                "observe": observe,
                "row": row,
            }
        )
    return verdicts, observes


def run_replay_round(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    rows = _load_pending_rows(project_root, ctx)
    verdicts: list[dict[str, Any]] = []
    observes: list[dict[str, Any]] = []
    replayed = False
    error = ""
    message_zh = ""
    if rows and _live_replay(ctx):
        try:
            verdicts, observes = _judge_rows(rows)
            replayed = True
        except Exception as exc:  # noqa: BLE001
            error, message_zh = _replay_bootstrap_failure(exc)
            doc = {
                "schema": "tg-replay-round/v1",
                "ok": False,
                "replayed": False,
                "error": error,
                "message_zh": message_zh,
                "detail": str(exc)[:400],
                "count": len(rows),
                "verdicts": [],
            }
            out = _receipt(project_root, ctx, "replay_round.yaml", doc)
            return {
                "ok": False,
                "engine": "replay_round",
                "error": error,
                "reason_code": error,
                "message_zh": message_zh,
                "artifact": out.as_posix(),
                "replayed": False,
                "count": len(rows),
            }
    if not verdicts:
        for i, row in enumerate(rows):
            observe = _observe_from_verdict(row, None)
            observes.append(observe)
            verdicts.append(
                {
                    "case_id": str(row.get("Testcase_Name") or f"case_{i}"),
                    "ok": False,
                    "tiling_key": row.get("tiling_key") or row.get("TilingKey") or "",
                    "reject": error or "NOT_RUN",
                    "judged": False,
                    "row": row,
                    "observe": observe,
                }
            )
    probed = False
    try:
        _text, fence = products.load_plan(_tg(project_root, ctx))
        from testcase_agent.coverage.probe import missing_probe_fields

        missing = missing_probe_fields(fence, observes) if _live_replay(ctx) else []
        if missing:
            inj = _try_inject_probes(project_root, ctx, missing)
            probed = bool(inj.get("ok"))
            if probed and rows and _live_replay(ctx):
                try:
                    verdicts, observes = _judge_rows(rows)
                    replayed = True
                except Exception:  # noqa: BLE001
                    pass
    except Exception:  # noqa: BLE001
        pass
    doc = {
        "schema": "tg-replay-round/v1",
        "ok": True,
        "replayed": replayed,
        "probed": probed,
        "error": error,
        "count": len(rows),
        "verdicts": verdicts,
    }
    out = _receipt(project_root, ctx, "replay_round.yaml", doc)
    cache = _replay_dir(project_root, ctx)
    cache.mkdir(parents=True, exist_ok=True)
    _dump_yaml(cache / "last_observes.yaml", {"observes": observes})
    return {"ok": True, "engine": "replay_round", "artifact": out.as_posix(), "replayed": replayed, "count": len(rows)}


def _probe_scope_files(project_root: Path, ctx: dict[str, Any]) -> list[str]:
    """Packet changed files first, then environment source_scope.file_paths."""
    paths: list[str] = []
    try:
        from ascendc_pilot.paths import agent_root

        arch = str(ctx.get("architecture") or "")
        rid = _run_id(ctx)
        packet_path = agent_root(project_root, arch) / "runs" / rid / "receipts" / "plan_scope_packet.yaml"
        packet = _load_yaml(packet_path) if packet_path.is_file() else {}
        for raw in packet.get("changed_files") or []:
            text = str(raw or "").strip()
            if text and text not in paths:
                paths.append(text)
    except Exception:  # noqa: BLE001
        pass
    try:
        from ascendc_pilot.environment_capabilities import source_scope_for_lease

        scope = source_scope_for_lease(project_root, run_id=_run_id(ctx))
        for raw in scope.get("allowed_source_files") or scope.get("file_paths") or []:
            text = str(raw or "").strip()
            if text and text not in paths:
                paths.append(text)
    except Exception:  # noqa: BLE001
        pass
    return paths


def _try_inject_probes(project_root: Path, ctx: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    from testcase_agent.coverage.probe import inject_probes
    from testcase_agent.closure.workspace import replay_runner

    try:
        from replay.bootstrap import ensure_runner, ops_sandbox_local, _ops_root
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": str(exc)}
    runner = replay_runner()
    boot = ensure_runner(runner, force_copy=True)
    if not boot.get("ok"):
        return {"ok": False, "error": boot.get("error")}
    raw_local = str(boot.get("ops_local") or "").strip()
    local = Path(raw_local) if raw_local else ops_sandbox_local(runner)
    try:
        original = _ops_root(runner)
        if local.resolve() == original.resolve():
            return {"ok": False, "error": "SANDBOX_COLLIDES_WITH_OPS"}
    except OSError:
        return {"ok": False, "error": "SANDBOX_OPS_UNRESOLVED"}
    if not local.is_dir():
        return {"ok": False, "error": "SANDBOX_OPS_MISSING"}
    injected = inject_probes(local, fields, scope=_probe_scope_files(project_root, ctx) or None)
    if injected.get("error") == "PROBE_AMBIGUOUS" or injected.get("ambiguous"):
        return {**injected, "ok": False, "error": "PROBE_AMBIGUOUS"}
    if injected.get("missing") and not injected.get("patched"):
        return {**injected, "ok": False, "error": "PROBE_UNTESTABLE"}
    rebuilt = ensure_runner(runner, rebuild=True)
    if not rebuilt.get("ok"):
        return {"ok": False, "error": rebuilt.get("error"), "inject": injected}
    return {"ok": True, "inject": injected, "rebuild": True}


def run_coverage_eval(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    from testcase_agent.coverage.eval import classify_eval_failure, evaluate_obligation
    from testcase_agent.coverage.contract import CASE_REFINABLE, CONTROL_GAP, OBSERVATION_GAP, PLAN_INVALID
    from testcase_agent.coverage.ledger import dump_worklog, ledger_closed, parse_worklog_fence, upsert_obligation

    tg = _tg(project_root, ctx)
    try:
        _text, fence = products.load_plan(tg)
    except products.ProductError as exc:
        return {"ok": False, "engine": "coverage_eval", "error": str(exc)}
    worklog = products.worklog_path(tg)
    ledger = parse_worklog_fence(worklog.read_text(encoding="utf-8") if worklog.is_file() else "")
    if not ledger.get("obligations"):
        return {"ok": False, "engine": "coverage_eval", "error": "ledger empty"}
    receipt_path = _action_dir(project_root, ctx, "replay_round").parent.parent / "receipts" / "replay_round.yaml"
    # receipts live under runs/{run_id}/receipts
    from ascendc_pilot.paths import agent_root

    arch = str(ctx.get("architecture") or "")
    receipt_path = agent_root(project_root, arch) / "runs" / _run_id(ctx) / "receipts" / "replay_round.yaml"
    replay_doc = _load_yaml(receipt_path) if receipt_path.is_file() else {}
    verdicts = replay_doc.get("verdicts") if isinstance(replay_doc, dict) else []
    observes = []
    for item in verdicts or []:
        if isinstance(item, dict) and isinstance(item.get("observe"), dict):
            observes.append(item["observe"])
    cache_obs = _load_yaml(_replay_dir(project_root, ctx) / "last_observes.yaml")
    if not observes and isinstance(cache_obs, dict):
        observes = [row for row in (cache_obs.get("observes") or []) if isinstance(row, dict)]
    seen = set(str(s) for s in (ledger.get("signatures") or []) if s)
    witnesses: list[dict[str, Any]] = []
    leak = False
    for observe in observes:
        rows = ledger.get("obligations") if isinstance(ledger.get("obligations"), dict) else {}
        for oid, obl in list(rows.items()):
            if not isinstance(obl, dict):
                continue
            status = str(obl.get("status") or "OPEN")
            if status not in {"OPEN", "MISS", "UNKNOWN"}:
                continue
            result = evaluate_obligation(obl, fence, observe, seen_signatures=seen)
            if result["status"] == "REDUNDANT":
                continue
            failure_class = classify_eval_failure(fence, obl, result, observe)
            upsert_obligation(
                ledger,
                oid,
                status=result["status"],
                signature=result.get("signature"),
                failure_class=failure_class,
            )
            if result["status"] == "CLOSED":
                seen.add(str(result.get("signature") or ""))
                row = (observe.get("case") if isinstance(observe.get("case"), dict) else {}) or {}
                witnesses.append({"obligation": oid, "row": row, "signature": result.get("signature")})
                obl["witness"] = {"row": row}
            if result["status"] == "GUARD_LEAK":
                leak = True
    isolation.assert_tg_write_path(worklog)
    worklog.write_text(dump_worklog(ledger), encoding="utf-8")
    cache = _replay_dir(project_root, ctx)
    cache.mkdir(parents=True, exist_ok=True)
    existing = _load_yaml(cache / "witnesses.yaml")
    kept = existing.get("witnesses") if isinstance(existing, dict) else []
    if not isinstance(kept, list):
        kept = []
    kept.extend(witnesses)
    _dump_yaml(cache / "witnesses.yaml", {"witnesses": kept})
    closed, problems = ledger_closed(ledger)
    reason = "GUARD_LEAK" if leak else _open_failure_reason(ledger)
    blocking = reason in {"GUARD_LEAK", "PLAN_INVALID", "HARNESS_CONTROL_GAP", "HARNESS_OBSERVATION_GAP"}
    return {
        "ok": not blocking,
        "engine": "coverage_eval",
        "closed": closed,
        "problems": problems,
        "guard_leak": leak,
        "reason_code": reason or (CASE_REFINABLE if not closed else ""),
        "artifact": worklog.as_posix(),
    }


def run_analyze_promote(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    from testcase_agent.coverage.ledger import merge_prose, open_ids, parse_worklog_fence
    from testcase_agent.coverage.contract import CASE_REFINABLE

    tg = _tg(project_root, ctx)
    text = _captured_text(project_root, ctx, "analyze_round")
    if not text.strip():
        text = _collect_staging_text(_action_dir(project_root, ctx, "analyze_round"), names=("worklog.md", "staging.md"))
    parsed = _parse_analyze_actions(text)
    extra_prose = ""
    if parsed:
        _receipt(project_root, ctx, "analyze_round.yaml", parsed)
    else:
        extra_prose = text
    path = products.worklog_path(tg)
    existing = path.read_text(encoding="utf-8") if path.is_file() else ""
    ledger = parse_worklog_fence(existing)
    if not ledger:
        return {"ok": False, "engine": "analyze_promote", "error": "missing worklog ledger"}
    isolation.assert_tg_write_path(path)
    path.write_text(merge_prose(existing, ledger, extra_prose=extra_prose), encoding="utf-8")
    remaining = open_ids(ledger)
    reason = _open_failure_reason(ledger)
    proof_requests = _extract_proof_requests(parsed)
    pending = bool(remaining) and bool(proof_requests)
    if not pending:
        _mark_not_applicable(
            project_root,
            ctx,
            ["source_proof", "proof_review", "proof_promote"],
            reason="no_proof_requests" if not proof_requests else "ledger_closed",
        )
    if remaining and not pending:
        blocking = reason in {"PLAN_INVALID", "HARNESS_CONTROL_GAP", "HARNESS_OBSERVATION_GAP"}
        code = reason if blocking else (reason or CASE_REFINABLE)
        if code == CASE_REFINABLE:
            code = "OPEN_REMAINING"
        return {
            "ok": False,
            "engine": "analyze_promote",
            "artifact": path.as_posix(),
            "open": remaining,
            "reason_code": code,
            "error": code,
        }
    return {
        "ok": True,
        "engine": "analyze_promote",
        "artifact": path.as_posix(),
        "open": remaining,
        "pending_proofs": len(proof_requests) if pending else 0,
    }


def run_proof_promote(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    from testcase_agent.coverage.contract import CASE_REFINABLE
    from testcase_agent.coverage.ledger import merge_prose, open_ids, parse_worklog_fence
    from testcase_agent.proof_promote import pair_items, promote

    items = pair_items(
        certificates=_parse_captured_docs(_captured_text(project_root, ctx, "source_proof")),
        reviews=_parse_captured_docs(_captured_text(project_root, ctx, "proof_review")),
    )
    tg = _tg(project_root, ctx)
    path = products.worklog_path(tg)
    existing = path.read_text(encoding="utf-8") if path.is_file() else ""
    ledger = parse_worklog_fence(existing)
    if not ledger:
        return {"ok": False, "engine": "proof_promote", "error": "missing worklog ledger"}
    out = promote(items=items, ledger=ledger)
    if not out.get("ok"):
        return {
            "ok": False,
            "engine": "proof_promote",
            "error": "PROOF_INVALID",
            "reason_code": "PROOF_INVALID",
            "errors": out.get("errors") or [],
        }
    isolation.assert_tg_write_path(path)
    path.write_text(merge_prose(existing, ledger), encoding="utf-8")
    remaining = open_ids(ledger)
    if remaining:
        reason = _open_failure_reason(ledger)
        blocking = reason in {"PLAN_INVALID", "HARNESS_CONTROL_GAP", "HARNESS_OBSERVATION_GAP"}
        code = reason if blocking else (reason or CASE_REFINABLE)
        if code == CASE_REFINABLE:
            code = "OPEN_REMAINING"
        return {
            "ok": False,
            "engine": "proof_promote",
            "artifact": path.as_posix(),
            "open": remaining,
            "applied": out.get("applied") or [],
            "reason_code": code,
            "error": code,
        }
    return {
        "ok": True,
        "engine": "proof_promote",
        "artifact": path.as_posix(),
        "open": remaining,
        "applied": out.get("applied") or [],
    }


def run_solve_certify(project_root: Path, ctx: dict[str, Any]) -> dict[str, Any]:
    from testcase_agent.coverage.ledger import ledger_closed, parse_worklog_fence

    tg = _tg(project_root, ctx)
    path = products.worklog_path(tg)
    if not path.is_file():
        return {"ok": False, "engine": "solve_certify", "error": "missing worklog.md"}
    text = path.read_text(encoding="utf-8")
    ledger = parse_worklog_fence(text)
    closed, problems = ledger_closed(ledger)
    if not closed:
        reason = _open_failure_reason(ledger) or "OPEN_NONEMPTY"
        if reason == "CASE_REFINABLE":
            reason = "OPEN_NONEMPTY"
        receipt = _receipt(
            project_root,
            ctx,
            "solve_certify.yaml",
            {"ok": False, "problems": problems, "worklog": path.as_posix(), "reason_code": reason},
        )
        return {
            "ok": False,
            "engine": "solve_certify",
            "problems": problems,
            "receipt": receipt.as_posix(),
            "reason_code": reason,
            "error": reason,
        }
    init_doc = products.load_init(tg)
    kind = str(init_doc.get("table_kind") or "csv")
    cases = products.cases_path(tg, kind)
    witnesses_doc = _load_yaml(_replay_dir(project_root, ctx) / "witnesses.yaml")
    rows = []
    for item in (witnesses_doc.get("witnesses") or []) if isinstance(witnesses_doc, dict) else []:
        if isinstance(item, dict) and isinstance(item.get("row"), dict):
            rows.append(item["row"])
    columns = [c["name"] if isinstance(c, dict) else str(c) for c in (init_doc.get("columns") or [])]
    if not columns and rows:
        columns = [str(k) for k in rows[0].keys()]
    isolation.assert_tg_write_path(cases)
    products.write_cases_table(cases, columns, rows, table_kind=kind)
    receipt = _receipt(
        project_root,
        ctx,
        "solve_certify.yaml",
        {"ok": True, "problems": [], "worklog": path.as_posix(), "cases": cases.as_posix(), "rows": len(rows)},
    )
    return {"ok": True, "engine": "solve_certify", "receipt": receipt.as_posix(), "rows": len(rows)}


def install(registry: dict[tuple[str, str], Any]) -> None:
    registry[("tg-init", "repo_scan")] = run_repo_scan
    registry[("tg-init", "bind_promote")] = run_bind_promote
    registry[("tg-init", "validate_init")] = run_validate_init
    registry[("tg-plan", "plan_precheck")] = run_plan_precheck
    registry[("tg-plan", "plan_promote")] = run_plan_promote
    registry[("tg-plan", "plan_validate")] = run_plan_validate
    registry[("tg-solve", "solve_precheck")] = run_solve_precheck
    registry[("tg-solve", "compile_obligations")] = run_compile_obligations
    registry[("tg-solve", "construct_promote")] = run_construct_promote
    registry[("tg-solve", "replay_round")] = run_replay_round
    registry[("tg-solve", "coverage_eval")] = run_coverage_eval
    registry[("tg-solve", "analyze_promote")] = run_analyze_promote
    registry[("tg-solve", "proof_promote")] = run_proof_promote
    registry[("tg-solve", "solve_certify")] = run_solve_certify
