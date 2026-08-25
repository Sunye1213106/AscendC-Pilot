# -*- coding: utf-8 -*-
"""Generic test-script repository intake for TG.

A test repo is optional. It is never operator-specific: the engine only
records filesystem facts (entry scripts, argparse flags, CSV headers, column profiles).
The agent reads those scripts against the CodeMap to learn how columns
become operator inputs, which flags mean precision vs performance, and
whether the scripts disagree with UO.

No test repo → generated cases use InputSemantics / knob defaults.
With a test repo → emitted rows must fill that repo's case schema so the
existing runner can consume them unchanged.
"""

from __future__ import annotations

import ast
import csv
import os
import re
from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any

SCHEMA = "tg-test-repo/v1"
_SKIP_DIR = {".git", "__pycache__", ".venv", "venv", "node_modules", ".mypy_cache"}
_CASE_HINTS = ("case", "csv", "xls", "xlsx", "table", "sheet")
_PRECISION_HINTS = ("only_grad", "precision", "compare", "atol", "rtol", "accuracy")
_PERF_HINTS = ("profiler", "profiling", "perf", "performance", "kernel_time")
_NOT_PRECISION_FLAGS = frozenset({"--golden-only", "--only_input_golden", "--no-save-golden"})
_ENABLE_NAMES = ("enable", "Enable", "ENABLE")
_PROFILE_TOPK = 16
_PROFILE_UNIQUE_CAP = 64
_INT_RE = re.compile(r"[+-]?\d+$")
_FLOAT_RE = re.compile(r"[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?$")


def default_contract(*, root: str = "", reason: str = "no_test_script_root") -> dict[str, Any]:
    return {
        "schema": SCHEMA,
        "kind": "default_input",
        "root": root,
        "entry": "",
        "case_arg": "",
        "modes": {"precision": [], "perf": []},
        "mode_candidates": [],
        "columns": [],
        "defaults": {},
        "corpus": [],
        "mapping": {},
        "findings": [],
        "reason": reason,
        "column_profile": {},
    }


def scan(root: str | Path | None) -> dict[str, Any]:
    """Structural inventory. No operator names, no LLM."""
    if not root:
        return {
            "schema": "tg-test-repo-inventory/v1",
            "kind": "default_input",
            "root": "",
            "entries": [],
            "flags": [],
            "tables": [],
            "canonical_call": {},
            "error": "",
        }
    path = Path(root).expanduser()
    try:
        path = path.resolve()
    except OSError:
        path = Path(root)
    if not path.is_dir():
        return {
            "schema": "tg-test-repo-inventory/v1",
            "kind": "missing",
            "root": str(path),
            "entries": [],
            "flags": [],
            "tables": [],
            "canonical_call": {},
            "error": f"test_script_root is not a directory: {path}",
        }
    entries: list[dict[str, Any]] = []
    flags: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    for file in _iter_files(path):
        rel = file.relative_to(path).as_posix()
        if file.suffix == ".py":
            parsed = _parse_python(file)
            if parsed["is_entry"] or parsed["flags"]:
                entries.append(
                    {
                        "path": rel,
                        "is_entry": parsed["is_entry"],
                        "relative_import": bool(parsed.get("relative_import")),
                        "imports": list(parsed.get("imports") or []),
                    }
                )
                for flag in parsed["flags"]:
                    flags.append({"path": rel, **flag})
        elif file.suffix.lower() in {".csv", ".tsv"}:
            header, sample, profile = _csv_table(file)
            if header:
                tables.append(
                    {
                        "path": rel,
                        "columns": header,
                        "sample": sample,
                        "kind": "csv",
                        "profile": profile,
                    }
                )
        elif file.suffix.lower() in {".xls", ".xlsx"}:
            header, sample, kind, err, profile = _xls_table(file)
            row = {"path": rel, "columns": header, "sample": sample, "kind": kind}
            if profile:
                row["profile"] = profile
            if err:
                row["error"] = err
            if header or err:
                tables.append(row)
    return {
        "schema": "tg-test-repo-inventory/v1",
        "kind": "script_repo",
        "root": str(path),
        "entries": entries,
        "flags": flags,
        "tables": tables,
        "canonical_call": scan_canonical_call(path),
        "error": "",
    }


def contract_from_inventory(
    inventory: dict[str, Any],
    *,
    host_fields: list[str] | None = None,
    key_dims: list[str] | None = None,
    knob_defaults: dict[str, Any] | None = None,
) -> dict[str, Any]:
    kind = str(inventory.get("kind") or "default_input")
    root = str(inventory.get("root") or "")
    if kind != "script_repo" or not root:
        doc = default_contract(root=root, reason=str(inventory.get("error") or "no_test_script_root"))
        if knob_defaults:
            doc["defaults"] = {str(k): _cell(v) for k, v in knob_defaults.items()}
            doc["columns"] = list(doc["defaults"])
        return doc

    tables = [t for t in (inventory.get("tables") or []) if isinstance(t, dict)]
    flags = [f for f in (inventory.get("flags") or []) if isinstance(f, dict)]
    entries = [e for e in (inventory.get("entries") or []) if isinstance(e, dict)]
    primary = _pick_table(tables)
    columns = list(primary.get("columns") or []) if primary else []
    sample = dict(primary.get("sample") or {}) if primary else {}
    profile = dict(primary.get("profile") or {}) if isinstance(primary.get("profile"), dict) else {}
    entry = _pick_entry(entries, flags)
    case_arg = _pick_case_arg(flags)
    entry_flags = _flags_for_entry(flags, entry, entries)
    modes = _pick_modes(entry_flags)
    mode_candidates = _mode_candidates(entry_flags)
    defaults = dict(sample)
    for name, value in (knob_defaults or {}).items():
        key = _match_column(columns, str(name)) or str(name)
        defaults.setdefault(key, _cell(value))

    findings = list(_cross_check(columns, host_fields or [], key_dims or []))
    if not entry:
        findings.append({"code": "missing_entry", "detail": "no runnable Python entry with argparse"})
    if not columns:
        findings.append({"code": "missing_schema", "detail": "no CSV/XLS case table with a header"})
    if inventory.get("error"):
        findings.append({"code": "scan_error", "detail": str(inventory["error"])})
    for table in tables:
        if table.get("error"):
            findings.append(
                {
                    "code": "unreadable_table",
                    "detail": f"{table.get('path')}: {table.get('error')}",
                }
            )
    owner = flag_owner(entry, entries, flags)
    if owner and owner != entry:
        findings.append(
            {
                "code": "entry_delegates_flags",
                "detail": f"{entry} is the runnable wrapper; argparse lives in {owner}",
            }
        )

    primary_table = str(primary.get("path") or "") if primary else ""
    return {
        "schema": SCHEMA,
        "kind": "script_repo",
        "root": root,
        "entry": entry,
        "flag_owner": owner,
        "case_arg": case_arg,
        "modes": modes,
        "mode_candidates": mode_candidates,
        "columns": columns,
        "defaults": defaults,
        "corpus": [str(t.get("path") or "") for t in tables if t.get("path")],
        "primary_table": primary_table,
        "table_kind": str(primary.get("kind") or "") if primary else "",
        "mapping": {},
        "findings": findings,
        "column_profile": profile,
        "canonical_call": inventory.get("canonical_call")
        if isinstance(inventory.get("canonical_call"), dict)
        else {},
        "reason": "",
    }


def fill_row(
    contract: dict[str, Any],
    knobs: dict[str, Any] | None = None,
    *,
    name: str = "",
    extra: dict[str, Any] | None = None,
) -> dict[str, str]:
    """Build one case the test repo can consume, or a default-input row."""
    columns = [str(c) for c in (contract.get("columns") or [])]
    defaults = dict(contract.get("defaults") or {})
    knobs = dict(knobs or {})
    extra = dict(extra or {})
    if not columns:
        columns = sorted({*_norm_keys(defaults), *_norm_keys(knobs), *_norm_keys(extra), "Testcase_Name"})
        if "tiling_key" not in {_fold(c) for c in columns}:
            columns.append("tiling_key")
    row: dict[str, str] = {}
    for col in columns:
        row[col] = _cell(
            _lookup(extra, col)
            or _lookup(knobs, col)
            or _lookup(defaults, col)
            or ("" if _fold(col) != _fold("Testcase_Name") else name)
        )
    if name:
        target = _match_column(columns, "Testcase_Name")
        if target:
            row[target] = name
    return row


def disable_illegal_row(contract: dict[str, Any], row: dict[str, str]) -> dict[str, str]:
    """P-ILLEGAL stays off-NPU: flip a generic enable column when present."""
    out = dict(row)
    col = _match_column(list(out) + list(contract.get("columns") or []), "enable")
    if col:
        out[col] = "disable"
    else:
        out["enable"] = "disable"
    return out


def _iter_files(root: Path) -> list[Path]:
    out: list[Path] = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in _SKIP_DIR and not d.startswith(".")]
        for name in filenames:
            path = Path(dirpath) / name
            if path.suffix.lower() in {".py", ".csv", ".tsv", ".xls", ".xlsx"}:
                out.append(path)
    return out


def _parse_python(path: Path) -> dict[str, Any]:
    try:
        tree = ast.parse(path.read_text(encoding="utf-8", errors="replace"))
    except (OSError, SyntaxError):
        return {"is_entry": False, "flags": [], "relative_import": False, "imports": []}
    flags: list[dict[str, Any]] = []
    imports: list[str] = []
    relative_import = False
    is_entry = path.name in {"main.py", "__main__.py"} or path.name.startswith("run_")
    for node in ast.walk(tree):
        if isinstance(node, ast.If) and _is_main_guard(node):
            is_entry = True
        if isinstance(node, ast.ImportFrom):
            # level > 0 is a package-relative import: the module cannot be run as a
            # plain script, only as `-m package.module` or via a wrapper.
            if (node.level or 0) > 0:
                relative_import = True
            if node.module:
                imports.append(str(node.module))
        elif isinstance(node, ast.Import):
            imports.extend(alias.name for alias in node.names if alias.name)
        if not isinstance(node, ast.Call):
            continue
        func_name = _call_name(node.func)
        if func_name != "add_argument" or not node.args:
            continue
        flag = _const_str(node.args[0])
        if not flag.startswith("-"):
            continue
        meta = {"flag": flag, "dest": "", "help": "", "takes_value": True, "choices": []}
        for kw in node.keywords:
            if kw.arg == "dest":
                meta["dest"] = _const_str(kw.value)
            elif kw.arg == "help":
                meta["help"] = _const_str(kw.value)
            elif kw.arg == "action" and _const_str(kw.value) in {"store_true", "store_false"}:
                meta["takes_value"] = False
            elif kw.arg == "choices":
                meta["choices"] = _const_list(kw.value)
        flags.append(meta)
    return {
        "is_entry": is_entry,
        "flags": flags,
        "relative_import": relative_import,
        "imports": sorted(set(imports)),
    }


def _is_main_guard(node: ast.If) -> bool:
    test = node.test
    if not isinstance(test, ast.Compare) or len(test.ops) != 1:
        return False
    if not isinstance(test.ops[0], ast.Eq):
        return False
    left = test.left
    return (
        isinstance(left, ast.Name)
        and left.id == "__name__"
        and bool(test.comparators)
        and _const_str(test.comparators[0]) == "__main__"
    )


def _call_name(func: ast.AST) -> str:
    if isinstance(func, ast.Name):
        return func.id
    if isinstance(func, ast.Attribute):
        return func.attr
    return ""


def _dotted_call(func: ast.AST) -> str:
    parts: list[str] = []
    cur: ast.AST | None = func
    while isinstance(cur, ast.Attribute):
        parts.append(cur.attr)
        cur = cur.value
    if isinstance(cur, ast.Name):
        parts.append(cur.id)
    parts.reverse()
    return ".".join(parts)


def scan_canonical_call(root: Path) -> dict[str, Any]:
    """Precision-path operator call with the most kwargs; skip profiler wrappers."""
    candidates: list[dict[str, Any]] = []
    path = Path(root)
    for file in _iter_files(path):
        if file.suffix.lower() != ".py":
            continue
        try:
            tree = ast.parse(file.read_text(encoding="utf-8", errors="replace"))
        except (OSError, SyntaxError):
            continue
        rel = file.relative_to(path).as_posix()
        if_blobs: list[str] = []
        with_blobs: list[str] = []

        class _Visitor(ast.NodeVisitor):
            def visit_If(self, node: ast.If) -> None:
                if_blobs.append(ast.dump(node.test).lower())
                self.generic_visit(node)
                if_blobs.pop()

            def visit_With(self, node: ast.With) -> None:
                names = [_dotted_call(item.context_expr).lower() for item in node.items]
                with_blobs.append(" ".join(names))
                self.generic_visit(node)
                with_blobs.pop()

            def visit_Call(self, node: ast.Call) -> None:
                api = _dotted_call(node.func)
                leaf = api.rsplit(".", 1)[-1].lower() if api else ""
                if not (leaf.startswith("npu_") or "aclnn" in leaf):
                    self.generic_visit(node)
                    return
                if "profiler" in api.lower():
                    self.generic_visit(node)
                    return
                ctx_if = " ".join(if_blobs)
                ctx_with = " ".join(with_blobs)
                profiler_ctx = "profiler" in ctx_with or (
                    "profiler" in ctx_if
                    and "only_grad" not in ctx_if
                    and "precision" not in ctx_if
                )
                precision_ctx = any(
                    token in ctx_if for token in ("only_grad", "precision", "auto_grad")
                )
                args = [str(kw.arg) for kw in node.keywords if kw.arg]
                kind = "aclnn" if "aclnn" in api.lower() else "pta"
                candidates.append(
                    {
                        "kind": kind,
                        "api": api,
                        "site": f"{rel}:{getattr(node, 'lineno', 0)}",
                        "args": args,
                        "score": len(node.args) + len(node.keywords),
                        "profiler": profiler_ctx,
                        "precision": precision_ctx,
                        "v2": "_v2" in leaf,
                        "grad": "grad" in leaf,
                    }
                )
                self.generic_visit(node)

        _Visitor().visit(tree)

    if not candidates:
        return {}
    usable = [row for row in candidates if not row["profiler"]] or candidates
    usable.sort(
        key=lambda row: (
            int(row["precision"]),
            int(row["v2"]),
            int(row["grad"]),
            int(row["score"]),
        ),
        reverse=True,
    )
    best = usable[0]
    return {
        "kind": str(best["kind"]),
        "api": str(best["api"]),
        "site": str(best["site"]),
        "args": list(best["args"]),
        "variant": "precision" if best["precision"] else "",
    }


def _const_str(node: ast.AST) -> str:
    if isinstance(node, ast.Constant):
        return str(node.value or "")
    return ""


def _const_list(node: ast.AST) -> list[str]:
    if isinstance(node, (ast.List, ast.Tuple, ast.Set)):
        return [item for item in (_const_str(elt) for elt in node.elts) if item]
    return []


def _csv_table(path: Path) -> tuple[list[str], dict[str, str], dict[str, Any]]:
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            header = [str(name) for name in (reader.fieldnames or []) if str(name).strip()]
            rows: list[dict[str, str]] = []
            for raw in reader:
                if not isinstance(raw, dict):
                    continue
                row = {str(k): _cell(v) for k, v in raw.items() if k}
                if _is_skipped_case_row(row, header):
                    continue
                rows.append(row)
            sample = dict(rows[0]) if rows else {}
            return header, sample, _profile_columns(header, rows)
    except OSError:
        return [], {}, {}


def _workbook_format(path: Path) -> str:
    """Detect the real workbook container, ignoring the file extension.

    Test tables are routinely saved as OOXML and renamed to ``.xls``; dispatching
    on the suffix sends those to xlrd, which refuses them.
    """
    try:
        with path.open("rb") as handle:
            magic = handle.read(8)
    except OSError:
        return "xlsx" if path.suffix.lower() == ".xlsx" else "xls"
    if magic.startswith(b"PK\x03\x04"):
        return "xlsx"
    if magic.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls"
    return "xlsx" if path.suffix.lower() == ".xlsx" else "xls"


def _xls_table(path: Path) -> tuple[list[str], dict[str, str], str, str, dict[str, Any]]:
    kind = _workbook_format(path)
    if kind == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            return [], {}, kind, "xlsx_reader_missing", {}
        try:
            # Pass bytes, not the path: openpyxl re-validates the suffix and would
            # reject an OOXML workbook that is named `.xls`.
            wb = load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            raw_header = next(it, ())
            header = [str(c).strip() for c in raw_header if str(c or "").strip()]
            rows: list[dict[str, str]] = []
            for sample_row in it:
                row = {
                    header[i]: _cell(sample_row[i] if i < len(sample_row) else "")
                    for i in range(len(header))
                }
                if _is_skipped_case_row(row, header):
                    continue
                rows.append(row)
            sample = dict(rows[0]) if rows else {}
            return header, sample, kind, "", _profile_columns(header, rows)
        except Exception as exc:  # noqa: BLE001
            return [], {}, kind, f"xlsx_read_failed:{type(exc).__name__}", {}
    try:
        import xlrd
    except ImportError:
        return [], {}, kind, "xls_reader_missing", {}
    try:
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        header = [
            str(sheet.cell_value(0, c)).strip()
            for c in range(sheet.ncols)
            if str(sheet.cell_value(0, c)).strip()
        ]
        rows = []
        for r in range(1, sheet.nrows):
            row = {
                header[c]: _cell(sheet.cell_value(r, c) if c < sheet.ncols else "")
                for c in range(len(header))
            }
            if _is_skipped_case_row(row, header):
                continue
            rows.append(row)
        sample = dict(rows[0]) if rows else {}
        return header, sample, kind, "", _profile_columns(header, rows)
    except Exception as exc:  # noqa: BLE001
        return [], {}, kind, f"xls_read_failed:{type(exc).__name__}", {}


def _is_skipped_case_row(row: dict[str, str], header: list[str]) -> bool:
    cells = [_cell(row.get(h, "")).strip() for h in header]
    if not any(cells):
        return True
    first = next((c for c in cells if c), "")
    if first.startswith("#"):
        return True
    name_col = _match_column(header, "Testcase_Name")
    if name_col and not _cell(row.get(name_col, "")).strip():
        return True
    return False


def _profile_columns(header: list[str], rows: list[dict[str, str]]) -> dict[str, Any]:
    n_rows = len(rows)
    columns: dict[str, Any] = {}
    for col in header:
        raw = [_cell(row.get(col, "")).strip() for row in rows]
        empty = sum(1 for v in raw if not v)
        nonempty = [v for v in raw if v]
        n_unique = len(set(nonempty))
        empty_rate = (empty / n_rows) if n_rows else 0.0
        inferred = _infer_column_type(nonempty, empty_rate)
        entry: dict[str, Any] = {
            "inferred_type": inferred,
            "n_unique": n_unique,
            "empty_rate": round(empty_rate, 4),
        }
        nums = _numeric_values(nonempty)
        if nums:
            entry["min"] = min(nums)
            entry["max"] = max(nums)
        topn = 4 if n_unique > _PROFILE_UNIQUE_CAP else _PROFILE_TOPK
        entry["topk"] = [
            {"value": value, "count": count}
            for value, count in Counter(nonempty).most_common(topn)
        ]
        if n_unique > _PROFILE_UNIQUE_CAP:
            entry["unique_truncated"] = True
        columns[col] = entry
    return {"n_rows": n_rows, "columns": columns}


def _infer_column_type(values: list[str], empty_rate: float) -> str:
    if empty_rate >= 0.8 or not values:
        return "empty-heavy"
    if all(_looks_list(v) for v in values):
        return "list-literal"
    nums = _numeric_values(values)
    if nums and all(isinstance(v, int) for v in nums):
        return "int"
    if nums:
        return "float"
    return "enum-string"


def _looks_list(value: str) -> bool:
    text = str(value).strip()
    return text.startswith("[") and text.endswith("]")


def _looks_int(value: str) -> bool:
    return bool(_INT_RE.fullmatch(str(value).strip()))


def _looks_float(value: str) -> bool:
    return bool(_FLOAT_RE.fullmatch(str(value).strip()))


def _numeric_values(values: list[str]) -> list[int] | list[float]:
    ints: list[int] = []
    floats: list[float] = []
    for value in values:
        text = str(value).strip()
        if _looks_int(text):
            ints.append(int(text, 10))
            floats.append(float(text))
        elif _looks_float(text):
            try:
                floats.append(float(text))
            except ValueError:
                return []
        else:
            return []
    if len(ints) == len(values):
        return ints
    if len(floats) == len(values):
        return floats
    return []


def _pick_table(tables: list[dict[str, Any]]) -> dict[str, Any]:
    if not tables:
        return {}
    scored = sorted(
        tables,
        key=lambda t: (
            -len(t.get("columns") or []),
            0 if "data/" in str(t.get("path") or "") else 1,
            str(t.get("path") or ""),
        ),
    )
    return scored[0]


def _module_names(path: str) -> set[str]:
    """Dotted module spellings a wrapper could use to import ``path``."""
    parts = Path(path).with_suffix("").parts
    if not parts:
        return set()
    dotted = ".".join(parts)
    names = {dotted, parts[-1]}
    if parts[-1] in {"main", "__main__"} and len(parts) > 1:
        names.add(".".join(parts[:-1]))
    return {n for n in names if n}


def _delegates_to(entry: dict[str, Any], candidates: list[str]) -> str:
    """Which flagged module this entry re-exports, if any."""
    imported = {str(name) for name in (entry.get("imports") or [])}
    for path in candidates:
        names = _module_names(path)
        if imported & names:
            return path
        # `from fag_test.main import main` records module `fag_test.main`;
        # also accept a parent-package import of the flagged module.
        if any(name.startswith(tuple(f"{n}." for n in names)) for name in imported):
            return path
    return ""


def flag_owner(entry: str, entries: list[dict[str, Any]], flags: list[dict[str, Any]]) -> str:
    """Path whose argparse flags apply when ``entry`` is executed."""
    flagged = [str(f.get("path") or "") for f in flags]
    ordered: list[str] = []
    for path in flagged:
        if path and path not in ordered:
            ordered.append(path)
    if entry in ordered:
        return entry
    row = next((e for e in entries if str(e.get("path") or "") == entry), None)
    if row:
        delegate = _delegates_to(row, ordered)
        if delegate:
            return delegate
    return entry


def _is_case_flag(row: dict[str, Any]) -> bool:
    blob = _flag_blob(row)
    return any(hint in blob for hint in _CASE_HINTS) and "cache" not in blob


def _pick_entry(entries: list[dict[str, Any]], flags: list[dict[str, Any]]) -> str:
    """Pick the script a human would actually run to execute cases.

    Two things beat "first file that has argparse": the entry must be executable as
    a script (a module using package-relative imports only runs via ``-m`` or via a
    wrapper), and its flags must include case selection, so a display/report tool
    with its own argparse does not win over the real driver.
    """
    by_path = {str(f.get("path") or ""): [] for f in flags}
    for row in flags:
        by_path.setdefault(str(row.get("path") or ""), []).append(row)
    flagged_order = [p for p in by_path if p]

    def rank(row: dict[str, Any]) -> tuple:
        path = str(row.get("path") or "")
        owner = path if path in by_path else _delegates_to(row, flagged_order)
        owned = by_path.get(owner) or []
        return (
            0 if not row.get("relative_import") else 1,
            0 if any(_is_case_flag(f) for f in owned) else 1,
            0 if owned else 1,
            len(Path(path).parts),
            0 if Path(path).name.startswith("run_") else 1,
            path,
        )

    candidates = [row for row in entries if row.get("is_entry")] or list(entries)
    if not candidates:
        return ""
    return str(sorted(candidates, key=rank)[0].get("path") or "")


def _pick_case_arg(flags: list[dict[str, Any]]) -> str:
    for row in flags:
        blob = _flag_blob(row)
        if any(hint in blob for hint in _CASE_HINTS) and "cache" not in blob:
            return str(row.get("flag") or "")
    return ""


def _flags_for_entry(
    flags: list[dict[str, Any]],
    entry: str,
    entries: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    if not entry:
        return list(flags)
    owner = flag_owner(entry, list(entries or []), flags)
    return [row for row in flags if str(row.get("path") or "") == owner]


def _mode_candidates(flags: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in flags:
        flag = str(row.get("flag") or "")
        if not flag or flag in _NOT_PRECISION_FLAGS:
            continue
        blob = _flag_blob(row)
        if any(hint in blob for hint in _CASE_HINTS) and "cache" not in blob:
            continue
        item: dict[str, Any] = {"flag": flag, "values": list(row.get("choices") or [])}
        out.append(item)
    return out


def _mode_blob(row: dict[str, Any]) -> str:
    """Flag text plus its declared choices.

    ``choices`` *are* the mode values, so a flag that documents nothing in help but
    declares ``choices=['only_grad', 'profiler']`` still identifies both modes. Kept
    separate from ``_flag_blob`` so case-selection detection is not widened.
    """
    return " ".join([_flag_blob(row), *(str(c or "").lower() for c in (row.get("choices") or []))])


def _pick_modes(flags: list[dict[str, Any]]) -> dict[str, list[str]]:
    precision: list[str] = []
    perf: list[str] = []
    for row in flags:
        blob = _mode_blob(row)
        flag = str(row.get("flag") or "")
        if not flag:
            continue
        if flag in _NOT_PRECISION_FLAGS:
            continue
        if any(hint in blob for hint in _PRECISION_HINTS):
            precision = _mode_argv(row, blob, _PRECISION_HINTS) or precision
        if any(hint in blob for hint in _PERF_HINTS):
            perf = _mode_argv(row, blob, _PERF_HINTS) or perf
    return {"precision": precision, "perf": perf}


def _mode_argv(row: dict[str, Any], blob: str, hints: tuple[str, ...]) -> list[str]:
    flag = str(row.get("flag") or "")
    if not row.get("takes_value"):
        return [flag]
    for hint in hints:
        if hint in blob:
            return [flag, hint]
    return [flag]


def _flag_blob(row: dict[str, Any]) -> str:
    return " ".join(
        str(row.get(key) or "").lower() for key in ("flag", "dest", "help")
    )


def _cross_check(columns: list[str], host_fields: list[str], key_dims: list[str]) -> list[dict[str, str]]:
    findings: list[dict[str, str]] = []
    known = {_fold(name): name for name in list(host_fields) + list(key_dims) if name}
    reserved = {_fold(name) for name in ("Testcase_Name", "enable", "tiling_key", * _ENABLE_NAMES)}
    for col in columns:
        folded = _fold(col)
        if folded in reserved or folded in known:
            continue
        if folded.startswith("actual") or folded.endswith("md5sum"):
            continue
        findings.append(
            {
                "code": "unmapped_column",
                "detail": f"case column {col!r} is not a host field or key dim; agent must map or flag a test-script gap",
            }
        )
    col_set = {_fold(c) for c in columns}
    for name in key_dims:
        if name and _fold(name) not in col_set and _fold(name) not in reserved:
            findings.append(
                {
                    "code": "missing_column",
                    "detail": f"UO key dim {name!r} has no case column; generated rows will use defaults",
                }
            )
    return findings


def _match_column(columns: list[str], name: str) -> str:
    want = _fold(name)
    for col in columns:
        if _fold(col) == want:
            return col
    return ""


def _lookup(payload: dict[str, Any], name: str) -> Any:
    if name in payload:
        return payload[name]
    want = _fold(name)
    for key, value in payload.items():
        if _fold(str(key)) == want:
            return value
    return None


def _norm_keys(payload: dict[str, Any]) -> list[str]:
    return [str(k) for k in payload]


def _fold(name: str) -> str:
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value)
